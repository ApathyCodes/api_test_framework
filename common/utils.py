"""
提供解析Excel 数据功能  使用openpyxl.load_Workbook

"""
import os
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Worksheet
from common.logger import logger


BaseDir = os.path.dirname(os.path.dirname(__file__))
dataDir = os.path.join(BaseDir, 'data')

# 封装类
class FileHandler():

    def __init__(self):
        self.data_dir = dataDir

    def do_Excel(self, excelpath: str, sheetname: str):
        """
        读取Excel文件，文件内容默认按行读取，读取结果存放到list中
        :param excelpath:
        :param sheetname:
        :return:
        """
        excel_file = os.path.join(self.data_dir, excelpath)

        # 加载Excel文件
        workbook = load_workbook(filename=excel_file)
        # 读取sheet
        worksheet: Worksheet = workbook[sheetname]

        # 获取到日志中sheet表
        logger.info(f'开始处理数据文件路径{excel_file}, sheet工作表{sheetname}')
        # 读取sheet表数据
        lists = []    # 定义空list
        for row in worksheet.iter_rows(
                min_col=1, min_row=2,
                max_row=worksheet.max_row,
                max_col=worksheet.max_column,
                values_only=True):

            print(row)
            lists.append(row)

        return lists


if __name__ == '__main__':
    fh = FileHandler()
    fh.do_Excel('testdata.xlsx','Sheet1')









